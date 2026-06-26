#!/usr/bin/env python3
"""
Gera a planilha de custo (modelo editável com fórmulas) do Nobre Bistrô.
3 abas: Custos (linhas por seção + subtotais + INSUMOS + perdas + CUSTO TOTAL) e
Precificação (custo-plus + 3 cenários + CMV). Edite CONFIG e rode.
Convenções: células azuis = inputs (preços a validar com fornecedor); amarelas = decisões.
Perdas e mão de obra do Breno SEMPRE entram (o tempo do chef é custo).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ====================== EDITE AQUI ======================
CONFIG = {
    "arquivo_saida": "Orcamento_NobreBistro.xlsx",
    "titulo": "Almoço de Natal (35 convidados · buffet)",
    "convidados": 35,
    "imposto": 0.10,          # alíquota efetiva do Simples (confirmar Anexo)
    "perda_pct": 0.10,        # perdas/sobra técnica sobre insumos
    "cenarios": [("Econômico", 0.12), ("Profissional", 0.22), ("Premium", 0.32)],
    # Seções de INSUMOS primeiro (entram no CMV e na base de perdas), depois as demais.
    # Linha = (item, qtd, unidade, preço_unit, observação)
    "secoes_insumos": ["ENTRADAS / SALADAS", "PRATOS", "SOBREMESAS"],
    "secao_operacional": "CUSTOS OPERACIONAIS",   # onde a linha de Perdas é inserida
    "secoes": [
        {"nome": "ENTRADAS / SALADAS", "itens": [
            ("Folhas / mix de verdes", 1.5, "kg", 30, ""),
            ("Frango (salpicão)", 2, "kg", 18, "estimativa"),
        ]},
        {"nome": "PRATOS", "itens": [
            ("Bacalhau dessalgado", 3.5, "kg", 110, "Natal sobe; validar"),
            ("Pernil suíno (c/ osso)", 7, "kg", 22, ""),
        ]},
        {"nome": "SOBREMESAS", "itens": [
            ("Pudim", 1, "lote", 85, ""),
            ("Cheesecake", 1, "lote", 130, ""),
            ("Torta de mousse de chocolate", 1, "lote", 110, ""),
        ]},
        {"nome": "DESCARTÁVEIS / EMBALAGENS", "itens": [
            ("Guardanapos premium", 1, "lote", 60, ""),
            ("Embalagens / transporte", 1, "lote", 120, ""),
        ]},
        {"nome": "CUSTOS OPERACIONAIS", "itens": [
            ("Gás", 1, "lote", 60, ""),
            ("Energia", 1, "lote", 50, ""),
            ("Transporte / logística", 1, "lote", 180, ""),
        ]},
        {"nome": "EQUIPE", "itens": [
            ("Copeira / auxiliar (feriado)", 2, "diária", 650, ""),
            ("Chef — produção (Breno)", 30, "h", 60, "horas × R$/h — NÃO esquecer"),
            ("Chef — no local (Breno)", 6, "h", 100, "feriado majorado"),
        ]},
        {"nome": "ESTRUTURA / LOCAÇÃO", "itens": [
            ("Rechauds", 5, "un", 35, ""),
            ("Decoração simples", 1, "lote", 120, ""),
        ]},
    ],
}
# ========================================================

F = "Arial"
navy = PatternFill("solid", fgColor="1F3864"); secfl = PatternFill("solid", fgColor="2E5496")
grey = PatternFill("solid", fgColor="D9E1F2"); agg = PatternFill("solid", fgColor="BDD7EE")
tot = PatternFill("solid", fgColor="FFE699"); grn = PatternFill("solid", fgColor="E2EFDA")
yel = PatternFill("solid", fgColor="FFFF00")
thin = Side(style="thin", color="BFBFBF"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
ctr = Alignment(horizontal="center", vertical="center")
lft = Alignment(horizontal="left", vertical="center", wrap_text=True)
rgt = Alignment(horizontal="right", vertical="center")
BRL = 'R$ #,##0.00;[Red](R$ #,##0.00);"-"'; PCT = '0.0%'
fnt = lambda **k: Font(name=F, **k)


def build():
    wb = Workbook()
    cs = wb.active; cs.title = "Custos"; cs.sheet_view.showGridLines = False
    for i, w in enumerate([46, 9, 11, 16, 15, 40], 1):
        cs.column_dimensions[get_column_letter(i)].width = w
    cs.merge_cells("A1:F1"); cs["A1"] = f"NOBRE BISTRÔ · Custos — {CONFIG['titulo']}"
    cs["A1"].font = fnt(bold=True, size=13, color="FFFFFF"); cs["A1"].fill = navy; cs["A1"].alignment = ctr
    cs.merge_cells("A2:F2")
    cs["A2"] = "Preços de REFERÊNCIA — substitua pelos custos reais do fornecedor (coluna Preço unit.). Azul = input; amarelo = validar."
    cs["A2"].font = fnt(italic=True, size=9, color="444444"); cs["A2"].alignment = lft
    for c, h in enumerate(["Item", "Qtd", "Un", "Preço unit. (R$)", "Custo (R$)", "Observação"], 1):
        x = cs.cell(row=4, column=c, value=h); x.font = fnt(bold=True, size=10, color="FFFFFF"); x.fill = secfl; x.alignment = ctr; x.border = bd

    r = 5
    insumos_subs, other_subs = [], []
    insumos_agg_written = False
    insumos_agg_row = None

    def section_header(r, name):
        cs.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        x = cs.cell(row=r, column=1, value=name); x.font = fnt(bold=True, size=10, color="FFFFFF"); x.fill = secfl; x.alignment = lft
        return r + 1

    def item_row(r, item, qtd, un, preco, obs):
        cs.cell(row=r, column=1, value=item).font = fnt(size=10); cs.cell(row=r, column=1).alignment = lft
        cs.cell(row=r, column=2, value=qtd).font = fnt(size=10, color="0000FF"); cs.cell(row=r, column=2).alignment = ctr
        cs.cell(row=r, column=3, value=un).font = fnt(size=10); cs.cell(row=r, column=3).alignment = ctr
        p = cs.cell(row=r, column=4, value=preco); p.font = fnt(size=10, color="0000FF"); p.alignment = rgt; p.number_format = BRL
        cs.cell(row=r, column=5, value=f"=B{r}*D{r}").number_format = BRL; cs.cell(row=r, column=5).alignment = rgt
        cs.cell(row=r, column=6, value=obs).font = fnt(size=8, color="666666"); cs.cell(row=r, column=6).alignment = lft
        for c in range(1, 7): cs.cell(row=r, column=c).border = bd
        return r + 1

    def subtotal_row(r, label, formula, fill=grey):
        cs.cell(row=r, column=1, value=label).font = fnt(bold=True, size=10)
        cs.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); cs.cell(row=r, column=1).alignment = rgt
        t = cs.cell(row=r, column=5, value=formula); t.font = fnt(bold=True, size=10); t.number_format = BRL; t.alignment = rgt; t.fill = fill
        cs.cell(row=r, column=1).fill = fill
        for c in range(1, 7): cs.cell(row=r, column=c).border = bd
        return r + 1

    for sec in CONFIG["secoes"]:
        is_insumo = sec["nome"] in CONFIG["secoes_insumos"]
        if not is_insumo and not insumos_agg_written:
            insumos_agg_row = r
            r = subtotal_row(r, "INSUMOS (Alimentos)", "=" + "+".join(insumos_subs), fill=agg)
            insumos_agg_written = True
        r = section_header(r, sec["nome"])
        first = r
        for (item, qtd, un, preco, obs) in sec["itens"]:
            r = item_row(r, item, qtd, un, preco, obs)
        # insere Perdas na seção operacional
        if sec["nome"] == CONFIG["secao_operacional"] and insumos_agg_row:
            cs.cell(row=r, column=1, value="Perdas / sobra técnica").font = fnt(size=10); cs.cell(row=r, column=1).alignment = lft
            q = cs.cell(row=r, column=2, value=CONFIG["perda_pct"]); q.font = fnt(size=10, color="0000FF"); q.fill = yel; q.number_format = PCT; q.alignment = ctr
            cs.cell(row=r, column=3, value="s/ insumos").font = fnt(size=10); cs.cell(row=r, column=3).alignment = ctr
            cs.cell(row=r, column=4, value="—").alignment = rgt
            cs.cell(row=r, column=5, value=f"=E{insumos_agg_row}*B{r}").number_format = BRL; cs.cell(row=r, column=5).alignment = rgt
            cs.cell(row=r, column=6, value="Padrão 8–12%").font = fnt(size=8, color="666666")
            for c in range(1, 7): cs.cell(row=r, column=c).border = bd
            r += 1
        last = r - 1
        sub_r = r
        r = subtotal_row(r, f"Subtotal {sec['nome'].title()}", f"=SUM(E{first}:E{last})")
        (insumos_subs if is_insumo else other_subs).append(f"E{sub_r}")

    r += 1
    cs.cell(row=r, column=1, value="CUSTO TOTAL OPERACIONAL").font = fnt(bold=True, size=11)
    cs.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4); cs.cell(row=r, column=1).alignment = rgt
    total_formula = "=E" + str(insumos_agg_row) + ("+" + "+".join(other_subs) if other_subs else "")
    g = cs.cell(row=r, column=5, value=total_formula); g.font = fnt(bold=True, size=11); g.number_format = BRL; g.alignment = rgt; g.fill = tot
    cs.cell(row=r, column=1).fill = tot
    for c in range(1, 7): cs.cell(row=r, column=c).border = bd
    custo_total_cell = f"Custos!E{r}"
    insumos_cell = f"Custos!E{insumos_agg_row}"

    # ---- Precificação ----
    ps = wb.create_sheet("Precificação"); ps.sheet_view.showGridLines = False
    for i, w in enumerate([30, 16, 18, 16, 16, 15, 14, 20], 1):
        ps.column_dimensions[get_column_letter(i)].width = w
    ps.merge_cells("A1:H1"); ps["A1"] = "NOBRE BISTRÔ · Formação de Preço e Cenários"
    ps["A1"].font = fnt(bold=True, size=13, color="FFFFFF"); ps["A1"].fill = navy; ps["A1"].alignment = ctr

    def kv(row, label, value, fmt=None, yf=False, b=False, link=False):
        ps.cell(row=row, column=1, value=label).font = fnt(bold=b, size=10); ps.cell(row=row, column=1).alignment = lft
        v = ps.cell(row=row, column=2, value=value); v.alignment = rgt
        v.font = fnt(size=10, color=("008000" if link else "0000FF"))
        if yf: v.fill = yel
        if fmt: v.number_format = fmt
        ps.cell(row=row, column=1).border = bd; v.border = bd

    kv(4, "Convidados", CONFIG["convidados"], fmt='0', yf=True)
    kv(5, "Imposto efetivo — Simples (%)", CONFIG["imposto"], fmt=PCT, yf=True)
    kv(7, "Custo total operacional (R$)", f"={custo_total_cell}", fmt=BRL, b=True, link=True)
    kv(8, "  Insumos (R$)", f"={insumos_cell}", fmt=BRL, link=True)
    kv(10, "  Custo por convidado (R$)", "=B7/B4", fmt=BRL)
    kv(11, "  CMV insumos / custo (%)", "=B8/B7", fmt=PCT)
    for c, h in enumerate(["Cenário", "Lucro líq. alvo", "Preço de venda", "Imposto", "Lucro líquido", "Markup", "CMV %", "Preço / convidado"], 1):
        x = ps.cell(row=13, column=c, value=h); x.font = fnt(bold=True, size=10, color="FFFFFF"); x.fill = secfl; x.alignment = ctr; x.border = bd
    for i, (nome, margem) in enumerate(CONFIG["cenarios"]):
        row = 14 + i
        ps.cell(row=row, column=1, value=nome).font = fnt(bold=True, size=10)
        m = ps.cell(row=row, column=2, value=margem); m.font = fnt(size=10, color="0000FF"); m.fill = yel; m.number_format = PCT; m.alignment = ctr
        ps.cell(row=row, column=3, value=f"=$B$7/(1-$B$5-B{row})").number_format = BRL
        ps.cell(row=row, column=4, value=f"=C{row}*$B$5").number_format = BRL
        ps.cell(row=row, column=5, value=f"=C{row}-$B$7-D{row}").number_format = BRL
        ps.cell(row=row, column=6, value=f"=C{row}/$B$7").number_format = '0.00"x"'
        ps.cell(row=row, column=7, value=f"=$B$8/C{row}").number_format = PCT
        ps.cell(row=row, column=8, value=f"=C{row}/$B$4").number_format = BRL
        for c in range(3, 9):
            ps.cell(row=row, column=c).alignment = rgt
            ps.cell(row=row, column=c).font = fnt(bold=(nome == "Profissional"), size=10)
        for c in range(1, 9):
            ps.cell(row=row, column=c).border = bd
            if nome == "Profissional": ps.cell(row=row, column=c).fill = grn

    wb.save(CONFIG["arquivo_saida"])
    print("Planilha gerada:", CONFIG["arquivo_saida"])


if __name__ == "__main__":
    build()
